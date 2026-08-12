from datetime import date, timedelta
from decimal import Decimal

from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Count, Q, F, Value, DecimalField
from django.db.models.functions import Coalesce
from django.shortcuts import render
from django.utils import timezone

from sales.models import Invoice, InvoiceItem
from customer_collections.models import Collection
from inventory.models import Batch, Inventory
from customers.models import Customer
from products.models import Product
from returns.models import Return

try:
    from consignments.models import Consignment
    HAS_CONSIGNMENTS = True
except Exception:
    HAS_CONSIGNMENTS = False

try:
    from manufacturing.models import ManufacturingOrder, Factory
    HAS_MANUFACTURING = True
except Exception:
    HAS_MANUFACTURING = False


def get_date_range(request):
    today = timezone.localdate()
    date_from = request.GET.get('date_from', today.replace(day=1).isoformat())
    date_to = request.GET.get('date_to', today.isoformat())
    try:
        date_from = date.fromisoformat(date_from)
        date_to = date.fromisoformat(date_to)
    except Exception:
        date_from = today.replace(day=1)
        date_to = today
    return date_from, date_to


@login_required
def reports_home(request):
    return render(request, 'reports/index.html', {})


@login_required
def index(request):
    return reports_home(request)


@login_required
def sales_report(request):
    company = request.company
    date_from, date_to = get_date_range(request)
    customer_id = request.GET.get('customer', '').strip()

    invoices = Invoice.objects.filter(
        company=company,
        invoice_date__gte=date_from,
        invoice_date__lte=date_to,
    ).select_related('customer')

    if customer_id:
        invoices = invoices.filter(customer_id=customer_id)

    totals = invoices.aggregate(
        total_sales=Coalesce(Sum('total_amount'), Value(0), output_field=DecimalField()),
        total_paid=Coalesce(Sum('amount_paid'), Value(0), output_field=DecimalField()),
        total_due=Coalesce(Sum('amount_due'), Value(0), output_field=DecimalField()),
        count=Count('id'),
    )

    by_customer = invoices.values(
        'customer__pharmacy_name'
    ).annotate(
        total=Coalesce(Sum('total_amount'), Value(0), output_field=DecimalField()),
        count=Count('id'),
        paid=Coalesce(Sum('amount_paid'), Value(0), output_field=DecimalField()),
        due=Coalesce(Sum('amount_due'), Value(0), output_field=DecimalField()),
    ).order_by('-total')[:20]

    by_product = []
    try:
        item_qs = InvoiceItem.objects.filter(
            invoice__company=company,
            invoice__invoice_date__gte=date_from,
            invoice__invoice_date__lte=date_to,
        )

        item_fields = [f.name for f in InvoiceItem._meta.fields]
        if 'is_bonus' in item_fields:
            item_qs = item_qs.filter(is_bonus=False)

        if 'line_total' in item_fields:
            by_product = item_qs.values(
                'product__product_name',
            ).annotate(
                total_qty=Coalesce(Sum('quantity'), Value(0), output_field=DecimalField()),
                total_value=Coalesce(Sum('line_total'), Value(0), output_field=DecimalField()),
            ).order_by('-total_value')[:20]
        else:
            by_product = item_qs.values(
                'product__product_name',
            ).annotate(
                total_qty=Coalesce(Sum('quantity'), Value(0), output_field=DecimalField()),
            ).order_by('-total_qty')[:20]
    except Exception:
        by_product = []

    customers = Customer.objects.filter(
        company=company,
        is_active=True
    ).order_by('pharmacy_name')

    return render(request, 'reports/sales_report.html', {
        'invoices': invoices.order_by('-invoice_date')[:50],
        'totals': totals,
        'by_customer': by_customer,
        'by_product': by_product,
        'customers': customers,
        'date_from': date_from,
        'date_to': date_to,
        'selected_customer': customer_id,
    })


@login_required
def collections_report(request):
    company = request.company
    date_from, date_to = get_date_range(request)
    customer_id = request.GET.get('customer', '').strip()
    method = request.GET.get('method', '').strip()

    collections = Collection.objects.filter(
        company=company,
        collection_date__gte=date_from,
        collection_date__lte=date_to,
    ).select_related('customer')

    if customer_id:
        collections = collections.filter(customer_id=customer_id)
    if method:
        collections = collections.filter(payment_method=method)

    totals = collections.aggregate(
        total=Coalesce(Sum('amount'), Value(0), output_field=DecimalField()),
        count=Count('id'),
    )

    by_method = collections.values('payment_method').annotate(
        total=Coalesce(Sum('amount'), Value(0), output_field=DecimalField()),
        count=Count('id'),
    ).order_by('-total')

    by_customer = collections.values(
        'customer__pharmacy_name'
    ).annotate(
        total=Coalesce(Sum('amount'), Value(0), output_field=DecimalField()),
        count=Count('id'),
    ).order_by('-total')[:20]

    customers = Customer.objects.filter(
        company=company,
        is_active=True
    ).order_by('pharmacy_name')

    return render(request, 'reports/collections_report.html', {
        'collections': collections.order_by('-collection_date')[:50],
        'totals': totals,
        'by_method': by_method,
        'by_customer': by_customer,
        'customers': customers,
        'date_from': date_from,
        'date_to': date_to,
        'selected_customer': customer_id,
        'selected_method': method,
    })


@login_required
def inventory_report(request):
    company = request.company
    today = timezone.localdate()

    product_id = request.GET.get('product', '').strip()
    stock_filter = request.GET.get('stock_filter', '').strip()

    rows = Inventory.objects.filter(
        company=company
    ).select_related('product', 'batch', 'warehouse').order_by(
        'product__product_name',
        'batch__expiry_date'
    )

    if product_id:
        rows = rows.filter(product_id=product_id)

    rows = list(rows)

    filtered_rows = []
    for row in rows:
        expiry_date = getattr(row.batch, 'expiry_date', None) if row.batch else None
        row.expiry_date = expiry_date
        row.days_until_expiry = (expiry_date - today).days if expiry_date else None

        include = True
        if stock_filter == 'near_expiry':
            include = bool(expiry_date and today <= expiry_date <= (today + timedelta(days=30)))
        elif stock_filter == 'expired':
            include = bool(expiry_date and expiry_date < today)
        elif stock_filter == 'low':
            min_level = getattr(row.product, 'min_stock_level', 0) or 0
            include = row.quantity_available <= min_level

        if include:
            filtered_rows.append(row)

    total_available = sum((r.quantity_available or 0) for r in filtered_rows)

    expiring_count = Inventory.objects.filter(
        company=company,
        batch__expiry_date__gte=today,
        batch__expiry_date__lte=today + timedelta(days=30)
    ).count()

    expired_count = Inventory.objects.filter(
        company=company,
        batch__expiry_date__lt=today
    ).count()

    products = Product.objects.filter(
        company=company,
        is_active=True
    ).order_by('product_name')

    return render(request, 'reports/inventory_report.html', {
        'rows': filtered_rows[:100],
        'total_available': total_available,
        'expiring_count': expiring_count,
        'expired_count': expired_count,
        'products': products,
        'date_today': today,
        'selected_product': product_id,
        'selected_stock_filter': stock_filter,
    })


@login_required
def customer_report(request):
    company = request.company
    search = request.GET.get('search', '').strip()

    customers = Customer.objects.filter(
        company=company,
        is_active=True
    ).annotate(
        due_balance=Coalesce(
            Sum('invoices__amount_due'),
            Value(0),
            output_field=DecimalField(max_digits=14, decimal_places=2)
        )
    ).order_by('pharmacy_name')

    if search:
        customers = customers.filter(
            Q(pharmacy_name__icontains=search) |
            Q(phone__icontains=search)
        )

    total_balance = customers.aggregate(
        t=Coalesce(
            Sum('due_balance'),
            Value(0),
            output_field=DecimalField(max_digits=14, decimal_places=2)
        )
    )['t'] or Decimal('0.00')

    over_limit = sum(1 for c in customers if (c.due_balance or 0) > (c.credit_limit or 0))

    return render(request, 'reports/customer_report.html', {
        'customers': customers[:100],
        'total_balance': total_balance,
        'over_limit': over_limit,
        'search': search,
    })


@login_required
def consignments_report(request):
    if not HAS_CONSIGNMENTS:
        return render(request, 'reports/index.html', {})

    company = request.company
    date_from, date_to = get_date_range(request)
    status_filter = request.GET.get('status', '').strip()
    customer_id = request.GET.get('customer', '').strip()

    consignments = Consignment.objects.filter(
        company=company,
        sent_date__gte=date_from,
        sent_date__lte=date_to,
    ).select_related('customer').prefetch_related('items')

    if status_filter:
        consignments = consignments.filter(status=status_filter)
    if customer_id:
        consignments = consignments.filter(customer_id=customer_id)

    total_sent_value = sum((c.total_sent_value or 0) for c in consignments)
    total_sold_value = sum((c.total_sold_value or 0) for c in consignments)
    total_returned_value = sum((c.total_returned_value or 0) for c in consignments)

    customers = Customer.objects.filter(
        company=company,
        is_active=True
    ).order_by('pharmacy_name')

    return render(request, 'reports/consignments_report.html', {
        'consignments': consignments,
        'total_sent_value': total_sent_value,
        'total_sold_value': total_sold_value,
        'total_returned_value': total_returned_value,
        'customers': customers,
        'status_choices': Consignment.STATUS_CHOICES,
        'date_from': date_from,
        'date_to': date_to,
        'selected_status': status_filter,
        'selected_customer': customer_id,
    })


@login_required
def manufacturing_report(request):
    if not HAS_MANUFACTURING:
        return render(request, 'reports/index.html', {})

    company = request.company
    status_filter = request.GET.get('status', '').strip()
    factory_id = request.GET.get('factory', '').strip()

    orders = ManufacturingOrder.objects.filter(
        company=company
    ).select_related('factory', 'product')

    if status_filter:
        orders = orders.filter(status=status_filter)
    if factory_id:
        orders = orders.filter(factory_id=factory_id)

    total_orders = orders.count()
    completed = orders.filter(status='completed').count()
    in_progress = orders.filter(status='in_progress').count()
    pending = orders.filter(status='pending').count()

    try:
        total_cost = orders.aggregate(t=Coalesce(Sum('total_cost'), Value(0), output_field=DecimalField()))['t'] or 0
    except Exception:
        total_cost = 0

    factories = Factory.objects.filter(
        company=company
    ).order_by('factory_name')

    return render(request, 'reports/manufacturing_report.html', {
        'orders': orders.order_by('-id')[:100],
        'total_orders': total_orders,
        'completed': completed,
        'in_progress': in_progress,
        'pending': pending,
        'total_cost': total_cost,
        'factories': factories,
        'status_choices': ManufacturingOrder.STATUS_CHOICES,
        'selected_status': status_filter,
        'selected_factory': factory_id,
    })


# ============================================
# Excel Export Views
# ============================================
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse


def get_company_info(company):
    try:
        from core.models import CompanyInfo
        return CompanyInfo.objects.filter(company=company).first()
    except Exception:
        return None


def excel_header_style(ws, company_info, title, company):
    """إضافة header للشركة في Excel"""
    ws.sheet_view.rightToLeft = True

    company_name = ''
    if company_info:
        company_name = company_info.company_name or ''
    if not company_name:
        company_name = getattr(company, 'name', 'PharmaFlow')

    # Row 1: Company name
    ws.merge_cells('A1:G1')
    ws['A1'] = company_name
    ws['A1'].font = Font(name='Cairo', size=16, bold=True, color='1a1d3a')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # Row 2: Report title
    ws.merge_cells('A2:G2')
    ws['A2'] = title
    ws['A2'].font = Font(name='Cairo', size=13, bold=True, color='4361ee')
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 25

    # Row 3: Info
    info_parts = []
    if company_info:
        if company_info.phone:
            info_parts.append(f"ت: {company_info.phone}")
        if company_info.tax_number:
            info_parts.append(f"ضريبي: {company_info.tax_number}")
        if company_info.address:
            info_parts.append(company_info.address)
    ws.merge_cells('A3:G3')
    ws['A3'] = ' | '.join(info_parts)
    ws['A3'].font = Font(name='Cairo', size=10, color='666666')
    ws['A3'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[3].height = 20

    # Separator
    ws.row_dimensions[4].height = 5

    return 5  # start data from row 5


def style_header_row(ws, row_num, headers):
    """تنسيق صف العناوين"""
    header_fill = PatternFill(start_color='4361EE', end_color='4361EE', fill_type='solid')
    header_font = Font(name='Cairo', bold=True, color='FFFFFF', size=11)
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='FFFFFF'),
        right=Side(style='thin', color='FFFFFF'),
        top=Side(style='thin', color='FFFFFF'),
        bottom=Side(style='thin', color='FFFFFF')
    )

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border

    ws.row_dimensions[row_num].height = 25


def style_data_row(ws, row_num, values, is_even=False):
    """تنسيق صف البيانات"""
    fill_color = 'F8F9FC' if is_even else 'FFFFFF'
    data_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
    data_font = Font(name='Cairo', size=10)
    data_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for col_num, value in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col_num, value=value)
        cell.fill = data_fill
        cell.font = data_font
        cell.alignment = data_align

    ws.row_dimensions[row_num].height = 20


@login_required
def export_sales_excel(request):
    company = request.company
    date_from, date_to = get_date_range(request)
    customer_id = request.GET.get('customer', '').strip()

    invoices = Invoice.objects.filter(
        company=company,
        invoice_date__gte=date_from,
        invoice_date__lte=date_to,
    ).select_related('customer').order_by('-invoice_date')

    if customer_id:
        invoices = invoices.filter(customer_id=customer_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'تقرير المبيعات'

    company_info = get_company_info(company)
    data_start = excel_header_style(
        ws, company_info,
        f'تقرير المبيعات | {date_from} إلى {date_to}',
        company
    )

    headers = ['#', 'رقم الفاتورة', 'العميل', 'التاريخ', 'الإجمالي', 'المدفوع', 'المتبقي', 'الحالة']
    style_header_row(ws, data_start, headers)

    status_map = {'issued': 'صادرة', 'partially_paid': 'جزئي', 'fully_paid': 'مدفوعة', 'cancelled': 'ملغية'}

    for i, inv in enumerate(invoices, 1):
        values = [
            i,
            inv.invoice_number,
            inv.customer.pharmacy_name,
            str(inv.invoice_date),
            float(inv.total_amount or 0),
            float(inv.amount_paid or 0),
            float(inv.amount_due or 0),
            status_map.get(inv.status, inv.status),
        ]
        style_data_row(ws, data_start + i, values, i % 2 == 0)

    # Totals row
    total_row = data_start + len(list(invoices)) + 1
    ws.merge_cells(f'A{total_row}:D{total_row}')
    ws[f'A{total_row}'] = 'الإجمالي'
    ws[f'A{total_row}'].font = Font(name='Cairo', bold=True, color='FFFFFF')
    ws[f'A{total_row}'].fill = PatternFill(start_color='1a1d3a', end_color='1a1d3a', fill_type='solid')
    ws[f'A{total_row}'].alignment = Alignment(horizontal='center')

    # Column widths
    col_widths = [5, 18, 25, 15, 15, 15, 15, 12]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="sales_report_{date_from}_{date_to}.xlsx"'
    wb.save(response)
    return response


@login_required
def export_collections_excel(request):
    company = request.company
    date_from, date_to = get_date_range(request)

    collections = Collection.objects.filter(
        company=company,
        collection_date__gte=date_from,
        collection_date__lte=date_to,
    ).select_related('customer').order_by('-collection_date')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'تقرير التحصيلات'

    company_info = get_company_info(company)
    data_start = excel_header_style(
        ws, company_info,
        f'تقرير التحصيلات | {date_from} إلى {date_to}',
        company
    )

    headers = ['#', 'العميل', 'التاريخ', 'المبلغ', 'طريقة الدفع', 'رقم الشيك', 'البنك']
    style_header_row(ws, data_start, headers)

    method_map = {'cash': 'نقدي', 'cheque': 'شيك', 'transfer': 'تحويل'}

    for i, col in enumerate(collections, 1):
        values = [
            i,
            col.customer.pharmacy_name,
            str(col.collection_date),
            float(col.amount or 0),
            method_map.get(col.payment_method, col.payment_method),
            col.cheque_number or '',
            col.bank_name or '',
        ]
        style_data_row(ws, data_start + i, values, i % 2 == 0)

    col_widths = [5, 25, 15, 15, 15, 15, 20]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="collections_report_{date_from}_{date_to}.xlsx"'
    wb.save(response)
    return response


@login_required
def export_inventory_excel(request):
    company = request.company

    rows = Inventory.objects.filter(
        company=company
    ).select_related('product', 'batch', 'warehouse').order_by(
        'product__product_name', 'batch__expiry_date'
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'تقرير المخزون'

    company_info = get_company_info(company)
    data_start = excel_header_style(ws, company_info, 'تقرير المخزون', company)

    headers = ['#', 'المنتج', 'المخزن', 'رقم الدفعة', 'تاريخ الانتهاء', 'المتاح', 'المحجوز', 'تحت التصريف']
    style_header_row(ws, data_start, headers)

    for i, row in enumerate(rows, 1):
        expiry = str(row.batch.expiry_date) if row.batch else ''
        values = [
            i,
            row.product.product_name,
            row.warehouse.warehouse_name if row.warehouse else '',
            row.batch.batch_number if row.batch else '',
            expiry,
            row.quantity_available or 0,
            row.quantity_reserved or 0,
            row.quantity_on_consignment or 0,
        ]
        style_data_row(ws, data_start + i, values, i % 2 == 0)

    col_widths = [5, 25, 15, 18, 15, 10, 10, 15]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="inventory_report.xlsx"'
    wb.save(response)
    return response


@login_required
def export_customers_excel(request):
    company = request.company

    customers = Customer.objects.filter(
        company=company,
        is_active=True
    ).order_by('pharmacy_name')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'تقرير العملاء'

    company_info = get_company_info(company)
    data_start = excel_header_style(ws, company_info, 'تقرير العملاء', company)

    headers = ['#', 'اسم العميل', 'الهاتف', 'المنطقة', 'حد الائتمان', 'الرصيد المستحق']
    style_header_row(ws, data_start, headers)

    for i, c in enumerate(customers, 1):
        due = sum(
            float(inv.amount_due or 0)
            for inv in c.invoices.filter(company=company)
        )
        values = [
            i,
            c.pharmacy_name,
            c.phone or '',
            c.area or '',
            float(c.credit_limit or 0),
            due,
        ]
        style_data_row(ws, data_start + i, values, i % 2 == 0)

    col_widths = [5, 30, 18, 20, 18, 18]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="customers_report.xlsx"'
    wb.save(response)
    return response


# ============================================
# كشف حساب عميل
# ============================================
@login_required
def customer_statement(request):
    company = request.company
    date_from, date_to = get_date_range(request)
    customer_id = request.GET.get('customer', '').strip()

    customers = Customer.objects.filter(
        company=company,
        is_active=True
    ).order_by('pharmacy_name')

    customer = None
    transactions = []
    opening_balance = Decimal('0.00')
    total_debit = Decimal('0.00')
    total_credit = Decimal('0.00')
    closing_balance = Decimal('0.00')

    if customer_id:
        try:
            customer = Customer.objects.get(pk=customer_id, company=company)

            # رصيد أول المدة
            sales_before = Invoice.objects.filter(
                company=company,
                customer=customer,
                invoice_date__lt=date_from
            ).aggregate(
                t=Coalesce(Sum('total_amount'), Value(0), output_field=DecimalField())
            )['t'] or Decimal('0.00')

            collections_before = Collection.objects.filter(
                company=company,
                customer=customer,
                collection_date__lt=date_from
            ).aggregate(
                t=Coalesce(Sum('amount'), Value(0), output_field=DecimalField())
            )['t'] or Decimal('0.00')

            try:
                returns_before = Return.objects.filter(
                    company=company,
                    customer=customer,
                    return_date__lt=date_from
                ).aggregate(
                    t=Coalesce(Sum('total_value'), Value(0), output_field=DecimalField())
                )['t'] or Decimal('0.00')
            except Exception:
                returns_before = Decimal('0.00')

            opening_balance = Decimal(str(sales_before)) - Decimal(str(collections_before)) - Decimal(str(returns_before))

            # فواتير البيع في الفترة
            invoices = Invoice.objects.filter(
                company=company,
                customer=customer,
                invoice_date__gte=date_from,
                invoice_date__lte=date_to,
            ).order_by('invoice_date', 'id')

            for inv in invoices:
                amt = Decimal(str(inv.total_amount or 0))
                transactions.append({
                    'date': inv.invoice_date,
                    'type': 'invoice',
                    'type_label': 'فاتورة بيع',
                    'reference': inv.invoice_number,
                    'debit': amt,
                    'credit': Decimal('0.00'),
                })
                total_debit += amt

            # التحصيلات في الفترة
            collections = Collection.objects.filter(
                company=company,
                customer=customer,
                collection_date__gte=date_from,
                collection_date__lte=date_to,
            ).order_by('collection_date', 'id')

            for col in collections:
                amt = Decimal(str(col.amount or 0))
                transactions.append({
                    'date': col.collection_date,
                    'type': 'collection',
                    'type_label': 'تحصيل',
                    'reference': col.receipt_number or f'RC-{col.id}',
                    'debit': Decimal('0.00'),
                    'credit': amt,
                })
                total_credit += amt

            # المرتجعات في الفترة
            try:
                returns = Return.objects.filter(
                    company=company,
                    customer=customer,
                    return_date__gte=date_from,
                    return_date__lte=date_to,
                ).order_by('return_date', 'id')

                for ret in returns:
                    amt = Decimal(str(ret.total_value or 0))
                    transactions.append({
                        'date': ret.return_date,
                        'type': 'return',
                        'type_label': 'مرتجع',
                        'reference': ret.return_number,
                        'debit': Decimal('0.00'),
                        'credit': amt,
                    })
                    total_credit += amt
            except Exception:
                pass

            # ترتيب حسب التاريخ
            transactions.sort(key=lambda x: x['date'])

            # حساب الرصيد الجاري
            running = opening_balance
            for tx in transactions:
                running += tx['debit'] - tx['credit']
                tx['balance'] = running

            closing_balance = running

        except Customer.DoesNotExist:
            customer_id = ''

    company_info = get_company_info(company)

    return render(request, 'reports/customer_statement.html', {
        'company_info': company_info,
        'customers': customers,
        'customer': customer,
        'transactions': transactions,
        'opening_balance': opening_balance,
        'total_debit': total_debit,
        'total_credit': total_credit,
        'closing_balance': closing_balance,
        'date_from': date_from,
        'date_to': date_to,
        'selected_customer': customer_id,
    })


@login_required
def export_customer_statement_excel(request):
    company = request.company
    date_from, date_to = get_date_range(request)
    customer_id = request.GET.get('customer', '').strip()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'كشف حساب'

    company_info = get_company_info(company)

    customer_name = 'عميل'
    try:
        customer = Customer.objects.get(pk=customer_id, company=company)
        customer_name = customer.pharmacy_name

        sales_before = Invoice.objects.filter(
            company=company, customer=customer,
            invoice_date__lt=date_from
        ).aggregate(
            t=Coalesce(Sum('total_amount'), Value(0), output_field=DecimalField())
        )['t'] or Decimal('0.00')

        collections_before = Collection.objects.filter(
            company=company, customer=customer,
            collection_date__lt=date_from
        ).aggregate(
            t=Coalesce(Sum('amount'), Value(0), output_field=DecimalField())
        )['t'] or Decimal('0.00')

        opening_balance = Decimal(str(sales_before)) - Decimal(str(collections_before))

        data_start = excel_header_style(
            ws, company_info,
            f'كشف حساب: {customer_name} | {date_from} إلى {date_to}',
            company
        )

        headers = ['#', 'التاريخ', 'البيان', 'المرجع', 'مدين', 'دائن', 'الرصيد']
        style_header_row(ws, data_start, headers)

        row_num = data_start + 1

        # رصيد أول المدة
        style_data_row(ws, row_num, [
            '', str(date_from), 'رصيد أول المدة', '',
            float(opening_balance) if opening_balance > 0 else 0,
            float(abs(opening_balance)) if opening_balance < 0 else 0,
            float(opening_balance)
        ], False)
        row_num += 1

        transactions = []

        for inv in Invoice.objects.filter(
            company=company, customer=customer,
            invoice_date__gte=date_from, invoice_date__lte=date_to
        ).order_by('invoice_date'):
            transactions.append({
                'date': inv.invoice_date,
                'label': 'فاتورة بيع',
                'ref': inv.invoice_number,
                'debit': Decimal(str(inv.total_amount or 0)),
                'credit': Decimal('0.00'),
            })

        for col in Collection.objects.filter(
            company=company, customer=customer,
            collection_date__gte=date_from, collection_date__lte=date_to
        ).order_by('collection_date'):
            transactions.append({
                'date': col.collection_date,
                'label': 'تحصيل',
                'ref': col.receipt_number or f'RC-{col.id}',
                'debit': Decimal('0.00'),
                'credit': Decimal(str(col.amount or 0)),
            })

        transactions.sort(key=lambda x: x['date'])
        running = opening_balance

        for i, tx in enumerate(transactions, 1):
            running += tx['debit'] - tx['credit']
            style_data_row(ws, row_num, [
                i,
                str(tx['date']),
                tx['label'],
                tx['ref'],
                float(tx['debit']),
                float(tx['credit']),
                float(running),
            ], i % 2 == 0)
            row_num += 1

        # إجمالي
        ws.merge_cells(f'A{row_num}:D{row_num}')
        ws[f'A{row_num}'] = 'الرصيد الختامي'

    except Exception:
        data_start = excel_header_style(ws, company_info, 'كشف حساب', company)
        style_header_row(ws, data_start, ['لم يتم اختيار عميل'])

    col_widths = [5, 14, 20, 20, 14, 14, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="statement_{customer_id}_{date_from}_{date_to}.xlsx"'
    wb.save(response)
    return response


# ============================================
# مجمل الربح
# ============================================
@login_required
def gross_profit_report(request):
    company = request.company
    date_from, date_to = get_date_range(request)

    # إجمالي المبيعات
    sales_total = Invoice.objects.filter(
        company=company,
        invoice_date__gte=date_from,
        invoice_date__lte=date_to,
    ).aggregate(
        t=Coalesce(Sum('total_amount'), Value(0), output_field=DecimalField())
    )['t'] or Decimal('0.00')

    # إجمالي المرتجعات
    try:
        returns_total = Return.objects.filter(
            company=company,
            return_date__gte=date_from,
            return_date__lte=date_to,
        ).aggregate(
            t=Coalesce(Sum('total_value'), Value(0), output_field=DecimalField())
        )['t'] or Decimal('0.00')
    except Exception:
        returns_total = Decimal('0.00')

    net_sales = Decimal(str(sales_total)) - Decimal(str(returns_total))

    # تكلفة البضاعة المباعة من بنود الفاتورة
    item_fields = [f.name for f in InvoiceItem._meta.fields]
    item_qs = InvoiceItem.objects.filter(
        invoice__company=company,
        invoice__invoice_date__gte=date_from,
        invoice__invoice_date__lte=date_to,
    ).select_related('product', 'batch')

    if 'is_bonus' in item_fields:
        item_qs = item_qs.filter(is_bonus=False)

    total_cost = Decimal('0.00')
    product_rows = {}

    for item in item_qs:
        qty = Decimal(str(item.quantity or 0))
        selling = Decimal(str(getattr(item, 'line_total', 0) or 0))

        # جيب التكلفة من الدفعة أو المنتج
        unit_cost = Decimal('0.00')
        if item.batch and getattr(item.batch, 'unit_cost', None):
            unit_cost = Decimal(str(item.batch.unit_cost))
        elif item.product and getattr(item.product, 'purchase_price', None):
            unit_cost = Decimal(str(item.product.purchase_price))

        line_cost = qty * unit_cost
        line_profit = selling - line_cost
        total_cost += line_cost

        pname = item.product.product_name if item.product else 'غير محدد'
        if pname not in product_rows:
            product_rows[pname] = {
                'name': pname,
                'qty': Decimal('0.00'),
                'selling': Decimal('0.00'),
                'cost': Decimal('0.00'),
                'profit': Decimal('0.00'),
            }
        product_rows[pname]['qty'] += qty
        product_rows[pname]['selling'] += selling
        product_rows[pname]['cost'] += line_cost
        product_rows[pname]['profit'] += line_profit

    gross_profit = net_sales - total_cost

    rows = sorted(product_rows.values(), key=lambda x: x['profit'], reverse=True)

    company_info = get_company_info(company)

    return render(request, 'reports/gross_profit_report.html', {
        'company_info': company_info,
        'date_from': date_from,
        'date_to': date_to,
        'sales_total': sales_total,
        'returns_total': returns_total,
        'net_sales': net_sales,
        'total_cost': total_cost,
        'gross_profit': gross_profit,
        'rows': rows,
    })


@login_required
def export_gross_profit_excel(request):
    company = request.company
    date_from, date_to = get_date_range(request)

    sales_total = Invoice.objects.filter(
        company=company,
        invoice_date__gte=date_from,
        invoice_date__lte=date_to,
    ).aggregate(
        t=Coalesce(Sum('total_amount'), Value(0), output_field=DecimalField())
    )['t'] or Decimal('0.00')

    try:
        returns_total = Return.objects.filter(
            company=company,
            return_date__gte=date_from,
            return_date__lte=date_to,
        ).aggregate(
            t=Coalesce(Sum('total_value'), Value(0), output_field=DecimalField())
        )['t'] or Decimal('0.00')
    except Exception:
        returns_total = Decimal('0.00')

    net_sales = Decimal(str(sales_total)) - Decimal(str(returns_total))

    item_fields = [f.name for f in InvoiceItem._meta.fields]
    item_qs = InvoiceItem.objects.filter(
        invoice__company=company,
        invoice__invoice_date__gte=date_from,
        invoice__invoice_date__lte=date_to,
    ).select_related('product', 'batch')

    if 'is_bonus' in item_fields:
        item_qs = item_qs.filter(is_bonus=False)

    total_cost = Decimal('0.00')
    product_rows = {}

    for item in item_qs:
        qty = Decimal(str(item.quantity or 0))
        selling = Decimal(str(getattr(item, 'line_total', 0) or 0))
        unit_cost = Decimal('0.00')
        if item.batch and getattr(item.batch, 'unit_cost', None):
            unit_cost = Decimal(str(item.batch.unit_cost))
        elif item.product and getattr(item.product, 'purchase_price', None):
            unit_cost = Decimal(str(item.product.purchase_price))
        line_cost = qty * unit_cost
        total_cost += line_cost
        pname = item.product.product_name if item.product else 'غير محدد'
        if pname not in product_rows:
            product_rows[pname] = {'name': pname, 'qty': Decimal('0.00'), 'selling': Decimal('0.00'), 'cost': Decimal('0.00'), 'profit': Decimal('0.00')}
        product_rows[pname]['qty'] += qty
        product_rows[pname]['selling'] += selling
        product_rows[pname]['cost'] += line_cost
        product_rows[pname]['profit'] += (selling - line_cost)

    gross_profit = net_sales - total_cost

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'مجمل الربح'

    company_info = get_company_info(company)
    data_start = excel_header_style(
        ws, company_info,
        f'تقرير مجمل الربح | {date_from} إلى {date_to}',
        company
    )

    # ملخص
    summary_start = data_start
    style_header_row(ws, summary_start, ['البند', 'المبلغ'])
    for i, row in enumerate([
        ['إجمالي المبيعات', float(sales_total)],
        ['المرتجعات', float(returns_total)],
        ['صافي المبيعات', float(net_sales)],
        ['تكلفة البضاعة المباعة', float(total_cost)],
        ['مجمل الربح', float(gross_profit)],
    ], 1):
        style_data_row(ws, summary_start + i, row, i % 2 == 0)

    # تفاصيل المنتجات
    detail_start = summary_start + 8
    style_header_row(ws, detail_start, ['#', 'المنتج', 'الكمية', 'المبيعات', 'التكلفة', 'الربح'])
    for i, row in enumerate(sorted(product_rows.values(), key=lambda x: x['profit'], reverse=True), 1):
        style_data_row(ws, detail_start + i, [
            i, row['name'], float(row['qty']),
            float(row['selling']), float(row['cost']), float(row['profit'])
        ], i % 2 == 0)

    for col, w in {'A': 6, 'B': 30, 'C': 12, 'D': 15, 'E': 15, 'F': 15}.items():
        ws.column_dimensions[col].width = w

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="gross_profit_{date_from}_{date_to}.xlsx"'
    wb.save(response)
    return response
